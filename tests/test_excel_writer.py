import pandas as pd
from openpyxl import load_workbook
from src.excel_writer import write_workbook


def test_workbook_has_all_sheets(tmp_path):
    summary = pd.DataFrame(
        {
            "cip": ["11.0701"],
            "program_name": ["CS"],
            "awards": [42],
            "wtd_median_wage_national": [105000.0],
        }
    )
    detail = pd.DataFrame(
        {"cip": ["11.0701"], "soc": ["15-1252"], "nat_median": [120000.0]}
    )
    crosswalk_used = pd.DataFrame({"cip": ["11.0701"], "soc": ["15-1252"]})
    methodology = [
        ("OEWS vintage", "May 2025 (released May 2026)"),
        ("IPEDS", "C2023_A; conferred Jul 2022-Jun 2023"),
    ]
    out = tmp_path / "wb.xlsx"
    write_workbook(summary, detail, crosswalk_used, methodology, out)

    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Detail", "Crosswalk Reference", "Methodology"]
    assert wb["Summary"]["A1"].value == "cip"
    # methodology label/value present
    meth_vals = [c.value for row in wb["Methodology"].iter_rows() for c in row]
    assert "OEWS vintage" in meth_vals


def test_summary_labels_vintages_and_openings_caveat(tmp_path):
    summary = pd.DataFrame(
        {
            "cip": ["11.0701"],
            "program_name": ["CS"],
            "wtd_growth_pct_national": [13.0],
            "total_annual_openings": [550.0],
        }
    )
    detail = pd.DataFrame(
        {
            "cip": ["11.0701"],
            "soc": ["15-1252"],
            "nat_growth_pct": [13.3],
            "dc_change_pct": [10.0],
        }
    )
    out = tmp_path / "wb.xlsx"
    write_workbook(
        summary, detail, pd.DataFrame({"cip": ["11.0701"]}), [("x", "y")], out
    )
    wb = load_workbook(out)
    summ_headers = [c.value for c in wb["Summary"][1]]
    assert "wtd_growth_pct_national_2024_34" in summ_headers
    assert "total_annual_openings_addressable" in summ_headers
    det_headers = [c.value for c in wb["Detail"][1]]
    assert "nat_growth_pct_2024_34" in det_headers
    assert "dc_change_pct_2022_32" in det_headers
    summ_cells = [c.value for row in wb["Summary"].iter_rows() for c in row if c.value]
    assert any("addressable-opportunity indicator" in str(v) for v in summ_cells)
